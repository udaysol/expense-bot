import pandas as pd
from database import get_connection

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def export_trip_to_excel(trip_id, trip_name, file_path):
    conn = get_connection()

    def style_sheet(ws, title_text):
        max_col = ws.max_column

        # Insert title row
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title_text
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_align

        # Style header row
        for col in range(1, max_col + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

            # Auto column width
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 16


    # ---- Load base data ----
    expenses = pd.read_sql_query(
        """
        SELECT id, description, paid_by, total_amount, split_type, timestamp
        FROM expenses
        WHERE trip_id=?
        ORDER BY timestamp
        """,
        conn,
        params=(trip_id,)
    )

    total_trip_expense = int(expenses["total_amount"].sum())

    splits = pd.read_sql_query(
        """
        SELECT s.expense_id, s.person, s.amount
        FROM expense_splits s
        JOIN expenses e ON s.expense_id = e.id
        WHERE e.trip_id=?
        """,
        conn,
        params=(trip_id,)
    )

    participants = pd.read_sql_query(
        "SELECT name FROM participants WHERE trip_id=?",
        conn,
        params=(trip_id,)
    )["name"].tolist()

    # ---- Build Expense Log Sheet ----
    rows = []

    for _, exp in expenses.iterrows():
        exp_splits = splits[splits["expense_id"] == exp["id"]]
        participant_count = len(exp_splits)

        row = {
            "Date": exp["timestamp"].split("T")[0],
            "Description": exp["description"],
            "Paid By": exp["paid_by"],
            "Total": exp["total_amount"],
            "Participants": participant_count,
            "Per Person": (
                exp["total_amount"] // participant_count
                if participant_count > 0 else 0
            ),
        }

        for p in participants:
            val = exp_splits.loc[exp_splits["person"] == p, "amount"]
            row[p] = int(val.iloc[0]) if not val.empty else 0

        rows.append(row)

    expense_log_df = pd.DataFrame(rows)

    total_row = {col: "" for col in expense_log_df.columns}
    total_row["Description"] = "TOTAL TRIP EXPENSE"
    total_row["Total"] = total_trip_expense

    expense_log_df = pd.concat(
        [expense_log_df, pd.DataFrame([total_row])],
        ignore_index=True
    )

    # ---- Person Summary Sheet ----
    paid = expenses.groupby("paid_by")["total_amount"].sum()
    owed = splits.groupby("person")["amount"].sum()

    summary_rows = []
    for p in participants:
        total_paid = int(paid.get(p, 0))
        total_owed = int(owed.get(p, 0))
        summary_rows.append([
            p,
            total_paid,
            total_owed,
            total_paid - total_owed
        ])

    summary_df = pd.DataFrame(
        summary_rows,
        columns=["Person", "Total Paid", "Total Owes", "Net (+Gets / -Pays)"]
    )

    if summary_df.empty:
        summary_df = pd.DataFrame(
            [["-", 0, 0, 0]],
            columns=["Person", "Total Paid", "Total Owes", "Net (+Gets / -Pays)"]
        )

    # ---- Settlement Sheet ----
    balances = summary_df.set_index("Person")["Net (+Gets / -Pays)"].to_dict()

    receivers = [[p, amt] for p, amt in balances.items() if amt > 0]
    payers = [[p, -amt] for p, amt in balances.items() if amt < 0]

    settlements = []
    i = j = 0

    while i < len(payers) and j < len(receivers):
        payer, owe = payers[i]
        receiver, recv = receivers[j]

        amt = min(owe, recv)
        settlements.append([payer, receiver, amt])

        payers[i][1] -= amt
        receivers[j][1] -= amt

        if payers[i][1] == 0:
            i += 1
        if receivers[j][1] == 0:
            j += 1

    if settlements:
        settlement_df = pd.DataFrame(
            settlements,
            columns=["From", "To", "Amount"]
        )
    else:
        settlement_df = pd.DataFrame(
            [["-", "-", 0]],
            columns=["From", "To", "Amount"]
        )


    conn.close()

    # Write to excel file
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
        expense_log_df.to_excel(writer, index=False, sheet_name="Expense Log")
        summary_df.to_excel(writer, index=False, sheet_name="Person Summary")
        settlement_df.to_excel(writer, index=False, sheet_name="Settlement")

        # ---- Styles ----
        header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)

        center_align = Alignment(horizontal="center")

        # ---- Apply styling SAFELY ----
        style_sheet(
            writer.sheets["Expense Log"],
            f"{trip_name.upper()} — EXPENSE LOG"
        )

        ws = writer.sheets["Expense Log"]
        last_row = ws.max_row

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
            )

        style_sheet(
            writer.sheets["Person Summary"],
            f"{trip_name.upper()} — PERSON SUMMARY"
        )

        style_sheet(
            writer.sheets["Settlement"],
            f"{trip_name.upper()} — FINAL SETTLEMENT"
        )

        # ---- Color Net Balances ----
        ws = writer.sheets["Person Summary"]
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=4)
            value = cell.value

            if isinstance(value, (int, float)):
                if value > 0:
                    cell.font = Font(color="006100", bold=True)
                elif value < 0:
                    cell.font = Font(color="9C0006", bold=True)
